"""Main Excel parser with format detection and logical unit assembly."""

from io import BytesIO

import openpyxl

from ..models.boq import ExcelFormat, LogicalUnit, ParsedBoQ, PricedLine
from .row_classifier import classify_row, ClassifiedRow, RowType


def parse_excel(content: bytes, filename: str) -> ParsedBoQ:
    """Parse an Excel BoQ file into structured logical units."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=False)

    # Find the BoQ sheet (skip cover pages and general conditions)
    sheet = _find_boq_sheet(wb)
    fmt = _detect_format(sheet)

    if fmt == ExcelFormat.FORMAT_A:
        rows = _classify_all_rows(sheet)
        units = _assemble_logical_units(rows)
        chapter_title = _extract_chapter_title(rows)
    else:
        # Format B stub — to be implemented
        rows = _classify_all_rows(sheet)
        units = _assemble_logical_units(rows)
        chapter_title = ""

    return ParsedBoQ(
        filename=filename,
        format_detected=fmt,
        sheet_name=sheet.title,
        chapter_title=chapter_title,
        units=units,
    )


def _find_boq_sheet(wb: openpyxl.Workbook):
    """Find the sheet containing the actual BoQ data."""
    # Prefer sheets with BOQ-like names
    for sheet in wb.worksheets:
        title_lower = sheet.title.lower()
        if any(kw in title_lower for kw in ("radovi", "troškovnik", "troskovnik", "stavk")):
            return sheet
    # Skip cover pages (NASLOVNICA) and general conditions
    for sheet in wb.worksheets:
        title_lower = sheet.title.lower()
        if title_lower not in ("naslovnica", "opći uvjeti", "opci uvjeti"):
            return sheet
    return wb.worksheets[0]


def _detect_format(sheet) -> ExcelFormat:
    """Detect whether this is Format A (Eurospin) or Format B (Kaufland)."""
    for row in sheet.iter_rows(min_row=1, max_row=30, values_only=False):
        values = [str(c.value).strip().lower() if c.value else "" for c in row]
        # Format A header: "stavka" in col B area, "jedinica mjere" nearby
        if any("stavka" in v for v in values) and any("jedinic" in v for v in values):
            # Check if data extends to column G (Format B) or stops at F (Format A)
            if len(row) >= 7 and row[6].value is not None:
                return ExcelFormat.FORMAT_B
            return ExcelFormat.FORMAT_A
        # Format B: code-like patterns in column A
        if values and "." in values[0] and values[0].replace(".", "").isdigit():
            parts = values[0].split(".")
            if any(len(p) >= 3 for p in parts):
                return ExcelFormat.FORMAT_B
    return ExcelFormat.FORMAT_A  # Default


def _classify_all_rows(sheet) -> list[ClassifiedRow]:
    """Classify every row in the sheet."""
    merged_ranges = list(sheet.merged_cells.ranges)
    results = []
    for row in sheet.iter_rows(min_row=1):
        classified = classify_row(row, merged_ranges)
        results.append(classified)
    return results


def _extract_chapter_title(rows: list[ClassifiedRow]) -> str:
    """Extract the main chapter title from classified rows."""
    for row in rows:
        if row.row_type == RowType.CHAPTER:
            return row.text
    return ""


def _assemble_logical_units(rows: list[ClassifiedRow]) -> list[LogicalUnit]:
    """Assemble classified rows into logical units.

    A logical unit is the deepest grouping that has priced sub-items beneath it.
    """
    units: list[LogicalUnit] = []
    current_chapter = ""
    current_section = ""
    current_subsection = ""
    current_item: LogicalUnit | None = None
    current_description_lines: list[str] = []

    for row in rows:
        if row.row_type == RowType.CHAPTER:
            current_chapter = row.item_number
        elif row.row_type == RowType.SECTION:
            current_section = row.item_number
            current_subsection = ""
        elif row.row_type == RowType.SUBSECTION:
            current_subsection = row.item_number
        elif row.row_type == RowType.WORK_ITEM:
            # Finalize previous item
            if current_item is not None:
                current_item.description = "\n".join(current_description_lines).strip()
                units.append(current_item)
            current_item = LogicalUnit(
                item_number=row.item_number,
                title=row.text,
                parent_section=current_subsection or current_section,
                parent_chapter=current_chapter or current_section,
                excel_start_row=row.excel_row,
                excel_end_row=row.excel_row,
            )
            current_description_lines = []
        elif row.row_type == RowType.DESCRIPTION:
            if current_item is not None:
                current_description_lines.append(row.text)
                current_item.excel_end_row = row.excel_row
        elif row.row_type == RowType.PRICED_LINE:
            if current_item is not None:
                current_item.priced_lines.append(
                    PricedLine(
                        item_number=row.item_number,
                        description=row.text,
                        unit=row.unit,
                        quantity=row.quantity,
                        unit_price=row.unit_price,
                        total=row.total,
                        excel_row=row.excel_row,
                    )
                )
                current_item.excel_end_row = row.excel_row

    # Finalize last item
    if current_item is not None:
        current_item.description = "\n".join(current_description_lines).strip()
        units.append(current_item)

    return units
