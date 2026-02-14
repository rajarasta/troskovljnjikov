from io import BytesIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

import openpyxl

from .upload import get_current_boq
from .output import get_output_data

router = APIRouter()


@router.post("/export")
async def export_excel():
    boq = get_current_boq()
    if not boq:
        raise HTTPException(404, "No BoQ uploaded yet")

    output_data = get_output_data()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Troskovnik"

    # Header row
    headers = ["RB", "STAVKA", "JEDINICA MJERE", "KOLIČINA", "JEDINIČNA CIJENA", "UKUPNA CIJENA"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    row_num = 2
    for unit in boq.units:
        # Unit title row
        ws.cell(row=row_num, column=1, value=unit.item_number)
        title_cell = ws.cell(row=row_num, column=2, value=unit.title)
        title_cell.font = openpyxl.styles.Font(bold=True)
        row_num += 1

        # Priced lines
        user_edits = {le.item_number: le for le in output_data.get(unit.id, [])}
        for pl in unit.priced_lines:
            ws.cell(row=row_num, column=1, value=pl.item_number)
            ws.cell(row=row_num, column=2, value=pl.description)
            ws.cell(row=row_num, column=3, value=pl.unit)
            ws.cell(row=row_num, column=4, value=pl.quantity)

            edit = user_edits.get(pl.item_number)
            price = edit.unit_price if edit else pl.unit_price
            ws.cell(row=row_num, column=5, value=price)
            if price is not None:
                ws.cell(row=row_num, column=6, value=round(pl.quantity * price, 2))
            row_num += 1

        row_num += 1  # Blank separator

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=output_{boq.filename}"},
    )
