"""File and folder indexing — top-level entry points."""

from __future__ import annotations

import os
import re
from datetime import datetime
from io import BytesIO
from typing import Any, Optional

import openpyxl

from .column_detection import find_best_header_row
from .constants import DATE_PATTERNS, NON_BOQ_SHEET_PATTERNS
from .item_extraction import extract_items_from_sheet


def resolve_file_date(file_name: str, file_path: str) -> tuple[datetime | None, str]:
    """Try to extract a date from the filename, falling back to file mtime.

    Returns:
        (date_or_none, source) where source is "filename", "file_mtime", or "manual".
    """
    for pattern, fmt in DATE_PATTERNS:
        m = re.search(pattern, file_name)
        if m:
            try:
                if fmt:
                    date_str = m.group(0).rstrip(".")
                    parsed = datetime.strptime(date_str, fmt)
                else:
                    # DDMMYYYY
                    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    parsed = datetime(y, mo, d)
                if 2000 <= parsed.year <= 2030:
                    return parsed, "filename"
            except (ValueError, OverflowError):
                continue

    try:
        mtime = os.path.getmtime(file_path)
        return datetime.fromtimestamp(mtime), "file_mtime"
    except OSError:
        pass

    return None, "manual"


def index_file(
    file_path: str,
    file_bytes: bytes,
    file_date: Optional[datetime] = None,
    llm_mapping: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Index a single BoQ file.

    Returns ``{"success": True/False, "file": {...}, "items": [...], "units": [...]}``.
    """
    try:
        file_name = os.path.basename(file_path)
        extension = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

        # Build lookup of LLM mappings by sheet name
        sheet_mappings: dict[str, Any] = {}
        if llm_mapping and llm_mapping.get("sheets"):
            for s in llm_mapping["sheets"]:
                if s.get("hasBOQData") and s.get("mapping"):
                    sheet_mappings[s["name"]] = s["mapping"]

        PREVIEW_ROWS = 10000  # capture all rows for native spreadsheet view

        items: list[dict[str, Any]] = []
        units: list[dict[str, Any]] = []
        sheet_infos: list[dict[str, Any]] = []
        raw_preview: dict[str, list[list[Any]]] = {}
        header_rows: dict[str, int] = {}
        column_mappings: dict[str, dict[str, Any]] = {}
        missing_price_count = 0
        missing_quantity_count = 0
        missing_unit_count = 0

        for sheet_name in wb.sheetnames:
            # If LLM mapping exists and this sheet is marked as non-BOQ, skip
            if llm_mapping and llm_mapping.get("sheets"):
                sheet_info_entry = next(
                    (s for s in llm_mapping["sheets"] if s["name"] == sheet_name),
                    None,
                )
                if sheet_info_entry and not sheet_info_entry.get("hasBOQData"):
                    sheet_infos.append({
                        "name": sheet_name,
                        "itemCount": 0,
                        "type": sheet_info_entry.get("type"),
                    })
                    continue

            # Auto-skip non-BoQ sheets by name when no LLM mapping
            if not llm_mapping:
                sheet_name_lower = sheet_name.lower().strip()
                if any(p in sheet_name_lower for p in NON_BOQ_SHEET_PATTERNS):
                    sheet_infos.append({
                        "name": sheet_name,
                        "itemCount": 0,
                        "type": "non_boq_auto",
                    })
                    continue

            ws = wb[sheet_name]

            # Capture first N raw rows for sheet preview
            preview_rows: list[list[Any]] = []
            for row_tuple in ws.iter_rows(min_row=1, max_row=PREVIEW_ROWS, max_col=21, values_only=True):
                preview_rows.append([
                    str(cell) if cell is not None else ""
                    for cell in row_tuple
                ])
            raw_preview[sheet_name] = preview_rows

            # Detect header row index for this sheet
            hr_index, _hr_map, hr_count = find_best_header_row(preview_rows)
            if hr_count >= 2:
                header_rows[sheet_name] = hr_index

            ext_mapping = sheet_mappings.get(sheet_name)
            sheet_result = extract_items_from_sheet(
                ws, file_name, file_path, sheet_name, file_date, ext_mapping,
            )
            items.extend(sheet_result["items"])
            units.extend(sheet_result["units"])

            # Collect column mapping for this sheet
            if sheet_result.get("columnMap"):
                column_mappings[sheet_name] = {
                    "columns": sheet_result["columnMap"],
                    "dataStartRow": sheet_result.get("dataStartRow", 0),
                }

            # Count missing data
            for item in sheet_result["items"]:
                if item.get("_missingPrice"):
                    missing_price_count += 1
                if item.get("_missingQuantity"):
                    missing_quantity_count += 1
                if item.get("_missingUnit"):
                    missing_unit_count += 1

            sheet_infos.append({
                "name": sheet_name,
                "itemCount": len(sheet_result["items"]),
                "type": "boq_data" if sheet_name in sheet_mappings else None,
                "usedLLMMapping": sheet_name in sheet_mappings,
            })

        wb.close()

        return {
            "success": True,
            "file": {
                "id": file_path,
                "fileName": file_name,
                "filePath": file_path,
                "fileType": extension,
                "date": file_date.isoformat() if file_date else None,
                "sheetCount": len(wb.sheetnames),
                "sheets": sheet_infos,
                "itemCount": len(items),
                "missingData": {
                    "prices": missing_price_count,
                    "quantities": missing_quantity_count,
                    "units": missing_unit_count,
                },
                "rawPreview": raw_preview,
                "headerRows": header_rows if header_rows else None,
                "columnMappings": column_mappings if column_mappings else None,
            },
            "items": items,
            "units": units,
        }
    except Exception as exc:
        return {
            "success": False,
            "error": str(exc),
            "file": {
                "id": file_path,
                "filePath": file_path,
                "error": str(exc),
            },
            "items": [],
            "units": [],
        }


def index_folder(folder_path: str) -> dict[str, Any]:
    """Index all BoQ files in a folder.

    Returns ``{"success": True, "folderPath": ..., "files": [...], "items": [...],
      "units": [...], "errors": [...], "stats": {...}}``.
    """
    supported_extensions = {"xlsx", "xls", "csv"}

    try:
        dir_entries = os.listdir(folder_path)
    except OSError as exc:
        raise RuntimeError(f"Failed to read directory: {exc}") from exc

    boq_files: list[str] = []
    for entry in dir_entries:
        ext = entry.rsplit(".", 1)[-1].lower() if "." in entry else ""
        if ext in supported_extensions:
            boq_files.append(entry)

    files: list[dict[str, Any]] = []
    items: list[dict[str, Any]] = []
    units: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for file_name in boq_files:
        full_path = os.path.join(folder_path, file_name).replace("\\", "/")
        try:
            with open(full_path, "rb") as f:
                file_bytes = f.read()

            mtime = os.path.getmtime(full_path)
            file_date = datetime.fromtimestamp(mtime)

            result = index_file(full_path, file_bytes, file_date)

            if result["success"]:
                files.append(result["file"])
                items.extend(result["items"])
                units.extend(result.get("units", []))
            else:
                errors.append({"filePath": full_path, "error": result.get("error")})
        except Exception as exc:
            errors.append({"filePath": full_path, "error": str(exc)})

    return {
        "success": True,
        "folderPath": folder_path,
        "files": files,
        "items": items,
        "units": units,
        "errors": errors,
        "stats": {
            "totalFiles": len(boq_files),
            "indexedFiles": len(files),
            "totalItems": len(items),
            "totalUnits": len(units),
            "errorCount": len(errors),
        },
    }
