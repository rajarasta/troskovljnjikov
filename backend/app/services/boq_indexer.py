"""
BoQ Indexer - Scans and indexes historical BoQ files from a folder.
Extracts items with descriptions, units, quantities, and prices.

Ported from boqIndexer.js (JavaScript/Electron) to Python/openpyxl.
"""

from __future__ import annotations

import os
import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.services.boq_hierarchy import build_hierarchy, group_into_units

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

COLUMN_PATTERNS: dict[str, list[str]] = {
    "itemNumber": [
        "r.br", "rbr", "r.b.", "rb", "redni broj", "redni",
        "item", "no", "no.", "number",
        "pozicija", "poz", "poz.",
        "šifra", "sifra", "oznaka",
        "kod", "code",
        "br.st", "br.st.", "broj stavke",
    ],
    "description": [
        "opis", "opis radova", "opis stavke",
        "description", "desc", "desc.",
        "naziv", "naziv stavke", "naziv radova",
        "stavka", "radovi", "rad",
        "kratki opis", "tekst stavke",
        "work", "item", "text", "tekst",
    ],
    "descriptionLong": [
        "dugi opis", "long description", "detailed description",
    ],
    "unit": [
        "jed", "jed.", "j.m.", "jm", "j.m", "jedinica", "jedinica mjere",
        "mjera", "mj", "mj.", "jed.mj", "jed.mj.",
        "unit", "uom", "u.o.m",
    ],
    "quantity": [
        "kol", "kol.", "količina", "kolicina", "količ.",
        "quantity", "qty", "qty.",
        "amount", "iznos kol", "broj",
    ],
    "unitPrice": [
        "jedinična cijena", "jedinicna cijena",
        "jed. cijena", "jed.cijena", "jed.cij", "jed.cij.",
        "cij", "cijena", "cij.",
        "j.c.", "jc", "j.c",
        "price", "unit price", "rate",
    ],
    "total": [
        "ukupna cijena", "ukupno", "ukup", "uk.", "ukup.",
        "total", "sum", "suma",
        "iznos", "vrijednost", "svega",
        "uc", "u.c", "u.c.",
    ],
}

MAX_HEADER_SEARCH_ROWS: int = 25

NON_BOQ_SHEET_PATTERNS: list[str] = [
    "naslovnica", "opći uvjeti", "opci uvjeti", "općiuvjeti",
    "export info", "cover", "general conditions",
]

KAUFLAND_CODE_RE = re.compile(r"^\d{3}(\.\d{3}(\.\d{5})?)?$")

KNOWN_UNITS: set[str] = {
    "m", "m2", "m²", "m3", "m³", "kom", "komad", "komplet", "kpl", "kg",
    "l", "lit", "set", "pauš", "paušal", "pau", "h", "sat", "dan",
    "cm", "mm", "km", "t", "ton", "%", "m1", "ml",
}


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

def detect_columns(header_row: list[Any]) -> tuple[dict[str, int], int]:
    """Detect column indices from a header row.

    Returns a tuple of (column_map, match_count).
    """
    column_map: dict[str, int] = {}
    match_count = 0
    used_columns: set[int] = set()

    normalized_headers = [
        {"index": i, "value": str(h or "").lower().strip()}
        for i, h in enumerate(header_row)
    ]

    for column_type, patterns in COLUMN_PATTERNS.items():
        if column_type in column_map:
            continue
        for header in normalized_headers:
            if column_type in column_map:
                break
            if not header["value"]:
                continue
            if header["index"] in used_columns:
                continue
            for pattern in patterns:
                if pattern in header["value"]:
                    column_map[column_type] = header["index"]
                    used_columns.add(header["index"])
                    match_count += 1
                    break

    return column_map, match_count


def find_best_header_row(rows: list[list[Any]]) -> tuple[int, dict[str, int], int]:
    """Search multiple rows for the best header row (most column pattern matches).

    Returns (header_row_index, column_map, match_count).
    """
    best_row = -1
    best_map: dict[str, int] = {}
    best_count = 0

    max_row = min(len(rows), MAX_HEADER_SEARCH_ROWS)
    for r in range(max_row):
        row = rows[r]
        if not row:
            continue
        col_map, match_count = detect_columns(row)
        if match_count > best_count:
            best_count = match_count
            best_map = col_map
            best_row = r

    return best_row, best_map, best_count


def detect_column_offset(rows: list[list[Any]], start_row: int) -> int:
    """Detect if column 0 is consistently empty (columns shifted right).

    Returns column offset (0 = no shift, 1 = shifted by 1).
    """
    empty_col0 = 0
    non_empty_col0 = 0
    check_count = min(20, len(rows) - start_row)

    for r in range(start_row, min(start_row + check_count, len(rows))):
        row = rows[r]
        if not row:
            continue
        val = str(row[0] if row[0] is not None else "").strip()
        if val == "":
            empty_col0 += 1
        else:
            non_empty_col0 += 1

    total = empty_col0 + non_empty_col0
    if total == 0:
        return 0

    if (empty_col0 > 0 and non_empty_col0 == 0) or (empty_col0 / total > 0.8):
        return 1
    return 0


def infer_columns_from_data(rows: list[list[Any]]) -> dict[str, Any] | None:
    """Try to infer column layout from data patterns when no header is found.

    Returns {"columnMap": dict, "dataStartRow": int} or None.
    """
    code_count = 0
    check_rows = min(15, len(rows))
    for r in range(check_rows):
        row = rows[r] if r < len(rows) else None
        val = str((row[0] if row and len(row) > 0 else None) or "").strip()
        if KAUFLAND_CODE_RE.match(val):
            code_count += 1

    if code_count < 3:
        return None

    # Kaufland format detected. Find the unit column by scanning data.
    col_count = max(
        (len(r) if r else 0) for r in rows[:check_rows]
    ) if check_rows > 0 else 0

    unit_scores = [0] * col_count

    for r in range(check_rows):
        row = rows[r] if r < len(rows) else None
        if not row:
            continue
        for c in range(3, min(len(row), col_count)):
            val = str(row[c] if row[c] is not None else "").lower().strip()
            if val in KNOWN_UNITS:
                unit_scores[c] += 1

    if not unit_scores:
        return None

    max_score = max(unit_scores)
    unit_col = unit_scores.index(max_score)

    if unit_col < 3 or unit_scores[unit_col] == 0:
        # Can't find unit column, use Kaufland default layout
        return {
            "columnMap": {
                "itemNumber": 0,
                "description": 2,
                "descriptionLong": 3,
                "unit": 5,
                "quantity": 4,
                "unitPrice": 6,
                "total": 7,
            },
            "dataStartRow": 0,
        }

    # Unit column found. Infer surrounding columns.
    column_map: dict[str, int | None] = {
        "itemNumber": 0,
        "description": 2,
        "descriptionLong": 3,
        "quantity": unit_col - 1,
        "unit": unit_col,
        "unitPrice": unit_col + 1,
        "total": unit_col + 2 if unit_col + 2 < col_count else None,
    }

    # Remove None values to match JS behavior (undefined keys not present)
    column_map_clean: dict[str, int] = {
        k: v for k, v in column_map.items() if v is not None
    }

    return {
        "columnMap": column_map_clean,
        "dataStartRow": 0,
    }


# ---------------------------------------------------------------------------
# Sheet / workbook helpers
# ---------------------------------------------------------------------------

def sheet_to_array(ws: Worksheet) -> list[list[Any]]:
    """Convert an openpyxl worksheet to a 2D list of cell values.

    Reads up to column U (index 21) to match the JS cap of ``Math.min(range.e.c, 20)``.
    """
    rows: list[list[Any]] = []
    for row_tuple in ws.iter_rows(min_row=1, max_col=21, values_only=True):
        rows.append([cell if cell is not None else "" for cell in row_tuple])
    return rows


# ---------------------------------------------------------------------------
# Item extraction
# ---------------------------------------------------------------------------

def extract_items_from_sheet(
    ws: Worksheet,
    file_name: str,
    file_path: str,
    sheet_name: str,
    file_date: Optional[datetime] = None,
    external_mapping: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Extract BoQ items from a worksheet.

    Parameters
    ----------
    ws : Worksheet
        An openpyxl worksheet object.
    file_name : str
        Source file name.
    file_path : str
        Source file path.
    sheet_name : str
        Sheet name.
    file_date : datetime, optional
        File modification date.
    external_mapping : dict, optional
        Optional LLM-detected column mapping with keys:
        ``headerRow``, ``dataStartRow``, ``columns``, ``skipPatterns``.

    Returns
    -------
    dict
        ``{"items": [...], "units": [...]}``
    """
    rows = sheet_to_array(ws)
    if len(rows) < 2:
        return {"items": [], "units": []}

    column_map: dict[str, int]
    data_start_row: int
    skip_patterns: list[str] = []

    if external_mapping:
        column_map = external_mapping["columns"]
        data_start_row = external_mapping.get("dataStartRow", 0)
        skip_patterns = external_mapping.get("skipPatterns", [])
    else:
        header_row_index, detected_map, match_count = find_best_header_row(rows)

        if match_count >= 3:
            column_map = detected_map
            data_start_row = header_row_index + 1

            # If partial header, try to fill missing columns from data patterns
            if column_map.get("unit") is None or column_map.get("quantity") is None:
                inferred = infer_columns_from_data(rows)
                if inferred:
                    inferred_map = inferred["columnMap"]
                    if column_map.get("unit") is None and inferred_map.get("unit") is not None:
                        column_map["unit"] = inferred_map["unit"]
                    if column_map.get("quantity") is None and inferred_map.get("quantity") is not None:
                        column_map["quantity"] = inferred_map["quantity"]
                    if column_map.get("unitPrice") is None and inferred_map.get("unitPrice") is not None:
                        column_map["unitPrice"] = inferred_map["unitPrice"]
                    if column_map.get("total") is None and inferred_map.get("total") is not None:
                        column_map["total"] = inferred_map["total"]
        else:
            inferred = infer_columns_from_data(rows)
            if inferred:
                column_map = inferred["columnMap"]
                data_start_row = inferred["dataStartRow"]
            else:
                offset = detect_column_offset(rows, 0)
                if offset > 0:
                    shifted_rows = [r[offset:] if r else r for r in rows]
                    shifted_row_idx, shifted_map, shifted_count = find_best_header_row(shifted_rows)

                    if shifted_count >= 3:
                        column_map = {
                            key: val + offset
                            for key, val in shifted_map.items()
                        }
                        data_start_row = shifted_row_idx + 1
                    else:
                        column_map = {
                            "itemNumber": 0 + offset,
                            "description": 1 + offset,
                            "unit": 2 + offset,
                            "quantity": 3 + offset,
                            "unitPrice": 4 + offset,
                            "total": 5 + offset,
                        }
                        data_start_row = 1
                else:
                    column_map = {
                        "itemNumber": 0,
                        "description": 1,
                        "unit": 2,
                        "quantity": 3,
                        "unitPrice": 4,
                        "total": 5,
                    }
                    data_start_row = 1

    # Handle dual-description columns: merge descriptionLong into description
    has_long_desc = "descriptionLong" in column_map

    if has_long_desc:
        desc_col = column_map["description"]
        long_col = column_map["descriptionLong"]
        for row in rows:
            if not row:
                continue
            short_desc = str(row[desc_col] if desc_col < len(row) else "").strip()
            long_desc = str(row[long_col] if long_col < len(row) else "").strip()
            if long_desc and short_desc:
                row[desc_col] = short_desc + " \u2014 " + long_desc
            elif long_desc and not short_desc:
                row[desc_col] = long_desc
        del column_map["descriptionLong"]

    project_name = extract_project_name(file_name, rows)

    # Use hierarchy-aware extraction
    hierarchy_items = build_hierarchy(rows, column_map, data_start_row, skip_patterns)

    # Group items into BoQ units
    boq_units = group_into_units(hierarchy_items, rows, column_map, file_path, sheet_name)

    # Build row -> unit_id lookup
    row_to_unit_id: dict[int, str] = {}
    for unit in boq_units:
        for item_id in unit.get("itemIds", []):
            # itemId format: "filePath:sheetName:row"
            parts = item_id.rsplit(":", 1)
            if len(parts) == 2:
                try:
                    row_num = int(parts[-1])
                    row_to_unit_id[row_num] = unit["id"]
                except ValueError:
                    pass

    date_str: str | None = None
    if file_date:
        date_str = file_date.strftime("%Y-%m-%d")

    items: list[dict[str, Any]] = []
    for hi in hierarchy_items:
        # Skip pure parent headers with no data
        if hi.get("isParent") and not hi.get("quantity") and not hi.get("unitPrice") and not hi.get("total"):
            continue

        item: dict[str, Any] = {
            "id": f"{file_path}:{sheet_name}:{hi['row']}",
            "fileId": file_path,
            "fileName": file_name,
            "filePath": file_path,
            "sheetName": sheet_name,
            "row": hi["row"],

            # BoQ data
            "itemNumber": hi.get("itemNumber"),
            "description": hi.get("description", ""),
            "fullDescription": hi.get("fullDescription"),
            "parentItemNumber": hi.get("parentItemNumber"),
            "unit": hi.get("unit"),
            "quantity": hi.get("quantity"),
            "unitPrice": hi.get("unitPrice"),
            "total": hi.get("total"),

            # Missing data flags
            "_missingPrice": "unitPrice" in column_map and not hi.get("unitPrice"),
            "_missingQuantity": "quantity" in column_map and not hi.get("quantity"),
            "_missingUnit": "unit" in column_map and not hi.get("unit"),

            # Unit reference
            "unitId": row_to_unit_id.get(hi["row"]),

            # Metadata
            "projectName": project_name,
            "date": date_str,
        }

        description = item.get("description", "") or ""
        full_description = item.get("fullDescription", "") or ""
        if len(description) >= 3 or len(full_description) >= 3:
            items.append(item)

    return {"items": items, "units": boq_units}


# ---------------------------------------------------------------------------
# Project name extraction
# ---------------------------------------------------------------------------

def extract_project_name(file_name: str, rows: list[list[Any]]) -> str:
    """Extract a project name from the filename or the first few rows."""
    name = re.sub(r"\.(xlsx?|csv|json)$", "", file_name, flags=re.IGNORECASE)
    name = re.sub(r"[-_]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()

    if re.match(r"^\d+$", name) or len(name) < 5:
        if rows and rows[0] and rows[0][0]:
            first_cell = str(rows[0][0])
            if len(first_cell) > 5:
                return first_cell[:100]

    return name


# ---------------------------------------------------------------------------
# File / folder indexing
# ---------------------------------------------------------------------------

def index_file(
    file_path: str,
    file_bytes: bytes,
    file_date: Optional[datetime] = None,
    llm_mapping: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    """Index a single BoQ file.

    Parameters
    ----------
    file_path : str
        Path to the file (used as identifier; the actual data comes from *file_bytes*).
    file_bytes : bytes
        Raw file content (xlsx/xls).
    file_date : datetime, optional
        File modification date.
    llm_mapping : dict, optional
        Optional LLM-detected mapping result with key ``sheets``:
        ``{"sheets": [{"name": ..., "hasBOQData": bool, "mapping": {...}}]}``.

    Returns
    -------
    dict
        ``{"success": True/False, "file": {...}, "items": [...], "units": [...]}``.
    """
    try:
        file_name = os.path.basename(file_path)
        extension = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

        # Build lookup of LLM mappings by sheet name
        sheet_mappings: dict[str, Any] = {}
        if llm_mapping and llm_mapping.get("sheets"):
            for s in llm_mapping["sheets"]:
                if s.get("hasBOQData") and s.get("mapping"):
                    sheet_mappings[s["name"]] = s["mapping"]

        items: list[dict[str, Any]] = []
        units: list[dict[str, Any]] = []
        sheet_infos: list[dict[str, Any]] = []
        missing_price_count = 0
        missing_quantity_count = 0
        missing_unit_count = 0

        for sheet_name in wb.sheetnames:
            # If LLM mapping exists and this sheet is marked as non-BOQ, skip
            if llm_mapping and llm_mapping.get("sheets"):
                sheet_info_entry = next(
                    (s for s in llm_mapping["sheets"] if s["name"] == sheet_name),
                    None,
                )
                if sheet_info_entry and not sheet_info_entry.get("hasBOQData"):
                    sheet_infos.append({
                        "name": sheet_name,
                        "itemCount": 0,
                        "type": sheet_info_entry.get("type"),
                    })
                    continue

            # Auto-skip non-BoQ sheets by name when no LLM mapping
            if not llm_mapping:
                sheet_name_lower = sheet_name.lower().strip()
                if any(p in sheet_name_lower for p in NON_BOQ_SHEET_PATTERNS):
                    sheet_infos.append({
                        "name": sheet_name,
                        "itemCount": 0,
                        "type": "non_boq_auto",
                    })
                    continue

            ws = wb[sheet_name]
            ext_mapping = sheet_mappings.get(sheet_name)
            sheet_result = extract_items_from_sheet(
                ws, file_name, file_path, sheet_name, file_date, ext_mapping,
            )
            items.extend(sheet_result["items"])
            units.extend(sheet_result["units"])

            # Count missing data
            for item in sheet_result["items"]:
                if item.get("_missingPrice"):
                    missing_price_count += 1
                if item.get("_missingQuantity"):
                    missing_quantity_count += 1
                if item.get("_missingUnit"):
                    missing_unit_count += 1

            sheet_infos.append({
                "name": sheet_name,
                "itemCount": len(sheet_result["items"]),
                "type": "boq_data" if sheet_name in sheet_mappings else None,
                "usedLLMMapping": sheet_name in sheet_mappings,
            })

        wb.close()

        return {
            "success": True,
            "file": {
                "id": file_path,
                "fileName": file_name,
                "filePath": file_path,
                "fileType": extension,
                "date": file_date.isoformat() if file_date else None,
                "sheetCount": len(wb.sheetnames),
                "sheets": sheet_infos,
                "itemCount": len(items),
                "missingData": {
                    "prices": missing_price_count,
                    "quantities": missing_quantity_count,
                    "units": missing_unit_count,
                },
            },
            "items": items,
            "units": units,
        }
    except Exception as exc:
        return {
            "success": False,
            "error": str(exc),
            "file": {
                "id": file_path,
                "filePath": file_path,
                "error": str(exc),
            },
            "items": [],
            "units": [],
        }


def index_folder(folder_path: str) -> dict[str, Any]:
    """Index all BoQ files in a folder (filesystem-based, no Electron API).

    Parameters
    ----------
    folder_path : str
        Path to the folder containing .xlsx / .xls / .csv files.

    Returns
    -------
    dict
        ``{"success": True, "folderPath": ..., "files": [...], "items": [...],
          "units": [...], "errors": [...], "stats": {...}}``.
    """
    supported_extensions = {"xlsx", "xls", "csv"}

    try:
        dir_entries = os.listdir(folder_path)
    except OSError as exc:
        raise RuntimeError(f"Failed to read directory: {exc}") from exc

    boq_files: list[str] = []
    for entry in dir_entries:
        ext = entry.rsplit(".", 1)[-1].lower() if "." in entry else ""
        if ext in supported_extensions:
            boq_files.append(entry)

    files: list[dict[str, Any]] = []
    items: list[dict[str, Any]] = []
    units: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for file_name in boq_files:
        full_path = os.path.join(folder_path, file_name).replace("\\", "/")
        try:
            with open(full_path, "rb") as f:
                file_bytes = f.read()

            mtime = os.path.getmtime(full_path)
            file_date = datetime.fromtimestamp(mtime)

            result = index_file(full_path, file_bytes, file_date)

            if result["success"]:
                files.append(result["file"])
                items.extend(result["items"])
                units.extend(result.get("units", []))
            else:
                errors.append({"filePath": full_path, "error": result.get("error")})
        except Exception as exc:
            errors.append({"filePath": full_path, "error": str(exc)})

    return {
        "success": True,
        "folderPath": folder_path,
        "files": files,
        "items": items,
        "units": units,
        "errors": errors,
        "stats": {
            "totalFiles": len(boq_files),
            "indexedFiles": len(files),
            "totalItems": len(items),
            "totalUnits": len(units),
            "errorCount": len(errors),
        },
    }
