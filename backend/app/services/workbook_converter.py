"""
Excel (.xlsx) to Univer IWorkbookData JSON converter.

Reads an Excel workbook via openpyxl and produces a dict matching Univer's
IWorkbookData schema, preserving cell values, formatting, merged ranges,
column widths, and row heights.
"""

from __future__ import annotations

import json
import uuid
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Approximate multiplier to convert openpyxl character-width units to pixels.
_CHAR_WIDTH_TO_PX = 8

# Mapping from openpyxl border style names to Univer numeric border styles.
_BORDER_STYLE_MAP: dict[str | None, int] = {
    "thin": 1,
    "medium": 2,
    "thick": 3,
    "dashed": 4,
    "dotted": 5,
    "double": 6,
    "hair": 7,
    "mediumDashed": 8,
    "dashDot": 9,
    "mediumDashDot": 10,
    "dashDotDot": 11,
    "mediumDashDotDot": 12,
    "slantDashDot": 13,
}

# Horizontal alignment mapping.
_H_ALIGN_MAP: dict[str | None, int] = {
    "left": 1,
    "center": 2,
    "right": 3,
    "fill": 1,
    "justify": 1,
    "centerContinuous": 2,
    "distributed": 1,
    "general": 1,
}

# Vertical alignment mapping.
_V_ALIGN_MAP: dict[str | None, int] = {
    "top": 1,
    "center": 2,
    "bottom": 3,
    "justify": 2,
    "distributed": 2,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_rgb(color_obj: Any) -> str | None:
    """Extract an #RRGGBB hex string from an openpyxl color object.

    Returns None when the colour cannot be resolved (theme colours without an
    rgb attribute, auto colours, or the sentinel "00000000").
    """
    if color_obj is None:
        return None
    try:
        rgb = color_obj.rgb
    except AttributeError:
        return None
    if rgb is None or not isinstance(rgb, str):
        return None
    # "00000000" is openpyxl's placeholder for auto / unset colours.
    if rgb == "00000000":
        return None
    # openpyxl stores ARGB as "FFRRGGBB"; strip the alpha channel.
    if len(rgb) == 8:
        return f"#{rgb[2:]}"
    if len(rgb) == 6:
        return f"#{rgb}"
    return None


def _cell_type(value: Any) -> int:
    """Return the Univer type code for a cell value.

    1 = string, 2 = number, 3 = boolean.
    """
    if isinstance(value, bool):
        return 3
    if isinstance(value, (int, float)):
        return 2
    return 1


def _cell_value(value: Any) -> str | int | float | bool:
    """Normalise a cell value for JSON serialisation."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def _build_border_side(side: Side | None) -> dict | None:
    """Convert an openpyxl Side to a Univer border-side dict."""
    if side is None or side.style is None:
        return None
    style_num = _BORDER_STYLE_MAP.get(side.style)
    if style_num is None:
        return None
    result: dict[str, Any] = {"s": style_num}
    color = _extract_rgb(side.color)
    if color is not None:
        result["cl"] = {"rgb": color}
    return result


def _build_style(cell: Cell) -> dict[str, Any] | None:
    """Build a Univer style dict for a single cell.

    Returns None if the cell carries no meaningful formatting.
    """
    style: dict[str, Any] = {}

    # --- Font properties ---
    font = cell.font
    if font is not None:
        if font.bold:
            style["bl"] = 1
        if font.italic:
            style["it"] = 1
        if font.size is not None:
            style["fs"] = font.size
        if font.name is not None:
            style["ff"] = font.name
        color = _extract_rgb(font.color)
        if color is not None:
            style["cl"] = {"rgb": color}

    # --- Fill (background) ---
    fill = cell.fill
    if fill is not None and fill.fill_type == "solid":
        bg = _extract_rgb(fill.start_color)
        if bg is not None:
            style["bg"] = {"rgb": bg}

    # --- Borders ---
    border: Border | None = cell.border
    if border is not None:
        bd: dict[str, Any] = {}
        for attr, key in (("top", "t"), ("bottom", "b"), ("left", "l"), ("right", "r")):
            side_dict = _build_border_side(getattr(border, attr, None))
            if side_dict is not None:
                bd[key] = side_dict
        if bd:
            style["bd"] = bd

    # --- Alignment ---
    alignment = cell.alignment
    if alignment is not None:
        ht = _H_ALIGN_MAP.get(alignment.horizontal)
        if ht is not None:
            style["ht"] = ht
        vt = _V_ALIGN_MAP.get(alignment.vertical)
        if vt is not None:
            style["vt"] = vt
        if alignment.wrap_text:
            style["tb"] = 3

    # --- Number format ---
    nf = cell.number_format
    if nf is not None and nf != "General":
        style["n"] = {"pattern": nf}

    return style if style else None


def _style_cache_key(style: dict[str, Any]) -> str:
    """Return a deterministic string key for a style dict (for dedup)."""
    return json.dumps(style, sort_keys=True)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def xlsx_to_workbook_data(file_path: str) -> dict:
    """Convert an Excel .xlsx file to a Univer IWorkbookData dict.

    Parameters
    ----------
    file_path:
        Path to the .xlsx file on disk.

    Returns
    -------
    dict
        A dictionary conforming to Univer's ``IWorkbookData`` schema.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    workbook_id = f"workbook-{uuid.uuid4().hex[:8]}"
    sheet_order: list[str] = []
    sheets: dict[str, Any] = {}
    # Global style cache shared across all sheets.
    style_cache: dict[str, str] = {}  # cache_key -> style_id
    styles: dict[str, dict] = {}      # style_id -> style dict
    style_counter = 0

    for ws in wb.worksheets:
        sheet_id = f"sheet-{uuid.uuid4().hex[:8]}"
        sheet_order.append(sheet_id)

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        row_count = max(max_row + 50, 1000)
        col_count = max(max_col + 10, 26)

        cell_data: dict[int, dict[int, dict]] = {}

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value is None:
                    continue

                r = cell.row - 1  # convert to 0-based
                c = cell.column - 1

                entry: dict[str, Any] = {
                    "v": _cell_value(cell.value),
                    "t": _cell_type(cell.value),
                }

                # Build and deduplicate style.
                style = _build_style(cell)
                if style is not None:
                    key = _style_cache_key(style)
                    if key not in style_cache:
                        sid = f"s{style_counter}"
                        style_counter += 1
                        style_cache[key] = sid
                        styles[sid] = style
                    entry["s"] = style_cache[key]

                cell_data.setdefault(r, {})[c] = entry

        # --- Auto-wrap description column ---
        # Detect description column: scan first few rows for "Stavka" header,
        # fall back to column B (index 1).
        desc_col = 1  # default: column B
        for scan_row in range(0, min(5, max_row)):
            row_cells = cell_data.get(scan_row, {})
            for ci, ce in row_cells.items():
                val = ce.get("v")
                if isinstance(val, str) and "stavka" in val.lower():
                    desc_col = ci
                    break

        # Inject tb=3 (wrap text) on all cells in the description column.
        for r_idx, row_cells in cell_data.items():
            entry = row_cells.get(desc_col)
            if entry is None:
                continue
            val = entry.get("v")
            if not isinstance(val, str) or len(val) < 10:
                continue
            existing_sid = entry.get("s")
            if existing_sid and existing_sid in styles:
                existing_style = styles[existing_sid]
                if existing_style.get("tb") == 3:
                    continue  # already wrapped
                new_style = {**existing_style, "tb": 3}
            else:
                new_style = {"tb": 3}
            new_key = _style_cache_key(new_style)
            if new_key not in style_cache:
                sid = f"s{style_counter}"
                style_counter += 1
                style_cache[new_key] = sid
                styles[sid] = new_style
            entry["s"] = style_cache[new_key]

        # --- Merged cells ---
        merge_data: list[dict[str, int]] = []
        for merged_range in ws.merged_cells.ranges:
            merge_data.append({
                "startRow": merged_range.min_row - 1,
                "endRow": merged_range.max_row - 1,
                "startColumn": merged_range.min_col - 1,
                "endColumn": merged_range.max_col - 1,
            })

        # --- Column widths ---
        column_data: dict[int, dict[str, Any]] = {}
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            dim = ws.column_dimensions.get(col_letter)
            if dim is not None and dim.width is not None:
                column_data[col_idx - 1] = {"w": round(dim.width * _CHAR_WIDTH_TO_PX)}

        # --- Row heights ---
        row_data: dict[int, dict[str, Any]] = {}
        for row_idx in range(1, max_row + 1):
            dim = ws.row_dimensions.get(row_idx)
            if dim is not None and dim.height is not None:
                row_data[row_idx - 1] = {"h": round(dim.height)}

        sheets[sheet_id] = {
            "id": sheet_id,
            "name": ws.title,
            "rowCount": row_count,
            "columnCount": col_count,
            "cellData": cell_data,
            "mergeData": merge_data,
            "columnData": column_data,
            "rowData": row_data,
            "defaultRowHeight": 20,
            "defaultColumnWidth": 80,
        }

    wb.close()

    return {
        "id": workbook_id,
        "sheetOrder": sheet_order,
        "sheets": sheets,
        "styles": styles,
    }
