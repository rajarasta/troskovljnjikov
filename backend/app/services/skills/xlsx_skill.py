"""
Excel (XLSX) Processing Skill - Create and analyze spreadsheets.

Provides tools for:
- Reading Excel files
- Creating Excel files
- Analyzing spreadsheet data
"""
import logging
from pathlib import Path
from typing import Optional, Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def read_excel_sheet(file_path: str, sheet_name: str = "Sheet1") -> dict:
    """
    Read an Excel sheet and return its data.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read

    Returns:
        Dictionary with sheet data
    """
    if not HAS_PANDAS:
        return {"error": "pandas not installed. Install with: pip install pandas"}

    try:
        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        df = pd.read_excel(path, sheet_name=sheet_name)
        return {
            "file": str(path),
            "sheet": sheet_name,
            "rows": len(df),
            "columns": list(df.columns),
            "data": df.to_dict(orient="records"),
            "success": True,
        }
    except Exception as e:
        logger.error(f"Error reading Excel sheet: {e}")
        return {"error": str(e), "success": False}


def list_excel_sheets(file_path: str) -> dict:
    """
    List all sheets in an Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary with sheet names and info
    """
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed. Install with: pip install openpyxl"}

    try:
        path = Path(file_path)
        if not path.exists():
            return {"error": f"File not found: {file_path}"}

        wb = load_workbook(path)
        sheets = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheets.append({
                "name": sheet,
                "rows": ws.max_row,
                "columns": ws.max_column,
            })

        return {
            "file": str(path),
            "sheet_count": len(sheets),
            "sheets": sheets,
            "success": True,
        }
    except Exception as e:
        logger.error(f"Error listing Excel sheets: {e}")
        return {"error": str(e), "success": False}


def create_excel_file(
    file_path: str,
    sheets: dict[str, list[list[Any]]],
    headers: Optional[dict[str, list[str]]] = None,
) -> dict:
    """
    Create a new Excel file with specified sheets and data.

    Args:
        file_path: Path where to save the Excel file
        sheets: Dictionary of {sheet_name: [rows of data]}
        headers: Optional dictionary of {sheet_name: [column headers]}

    Returns:
        Dictionary with creation status
    """
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed"}

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(title=sheet_name)

            # Add headers if provided
            if headers and sheet_name in headers:
                for col_idx, header in enumerate(headers[sheet_name], 1):
                    ws.cell(row=1, column=col_idx, value=header)
                start_row = 2
            else:
                start_row = 1

            # Add data rows
            for row_idx, row in enumerate(rows, start_row):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)
        return {
            "file": str(file_path),
            "sheets_created": len(sheets),
            "success": True,
        }
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        return {"error": str(e), "success": False}


# Skill registration info
SKILL_INFO = {
    "name": "xlsx_handler",
    "description": "Create and analyze Excel spreadsheets",
    "functions": [
        {
            "name": "read_excel_sheet",
            "description": "Read data from an Excel sheet",
            "parameters": {
                "file_path": "Path to the Excel file",
                "sheet_name": "Name of sheet to read (default: Sheet1)",
            },
        },
        {
            "name": "list_excel_sheets",
            "description": "List all sheets in an Excel file",
            "parameters": {
                "file_path": "Path to the Excel file",
            },
        },
        {
            "name": "create_excel_file",
            "description": "Create a new Excel file with specified data",
            "parameters": {
                "file_path": "Path where to save the file",
                "sheets": "Dictionary of sheet names and data",
                "headers": "Optional column headers per sheet",
            },
        },
    ],
}
