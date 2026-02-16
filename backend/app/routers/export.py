from __future__ import annotations

import io
import os
import re
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook as XLWorkbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.boq import BoQFile, BoQItem, ItemStatusEnum
from app.models.preset import Preset
from app.services.column_registry import ColumnDef, resolve_columns

router = APIRouter()

# ── Status colour map ────────────────────────────────────────────────

_STATUS_COLORS: dict[str, str] = {
    "pending": "FFF9C4",
    "accepted": "C8E6C9",
    "refused": "FFCDD2",
    "negotiated": "BBDEFB",
    "expired": "E0E0E0",
}

_STATUS_LABELS: dict[str, str] = {
    "pending": "Na \u010dekanju",
    "accepted": "Prihva\u0107eno",
    "refused": "Odbijeno",
    "negotiated": "Dogovoreno",
    "expired": "Isteklo",
}


def _get_file_with_items(
    file_id: str, db: Session
) -> tuple[BoQFile, list[BoQItem]]:
    boq_file = db.query(BoQFile).filter(BoQFile.id == file_id).first()
    if not boq_file:
        raise HTTPException(status_code=404, detail="File not found")
    items = (
        db.query(BoQItem)
        .filter(BoQItem.file_id == file_id)
        .order_by(BoQItem.row)
        .all()
    )
    return boq_file, items


# ── Helpers ──────────────────────────────────────────────────────────


def _sanitize(name: str, max_len: int) -> str:
    """Replace non-alphanumeric chars with underscore and truncate."""
    cleaned = re.sub(r"[^\w]+", "_", name or "export").strip("_")
    return cleaned[:max_len]


def _get_cell_value(item: BoQItem, col: ColumnDef):
    """Extract the display value for a column from a BoQItem."""
    key = col.key

    # Status group columns come from the status_record relationship
    if key == "status":
        sr = item.status_record
        return sr.status.value if sr else ""
    if key == "updated_at":
        sr = item.status_record
        if sr and sr.updated_at:
            return sr.updated_at.strftime("%d.%m.%Y.")
        return ""

    # Drawing is handled specially (image embedding) — return the path
    if key == "drawing":
        return getattr(item, "drawing_path", None) or ""

    # Everything else is a direct attribute
    return getattr(item, key, "") or ("" if col.col_type != "float" else 0)


def _is_composite_parent(item: BoQItem, items: list[BoQItem]) -> bool:
    """True when item has children (composite unit parent) but no own price."""
    if item.item_type == "section_header":
        return False
    if item.unit_price or item.total:
        return False
    # Check if any later item references this item as parent
    if item.item_number:
        return any(
            other.parent_item_number == item.item_number for other in items
        )
    return False


# ── XLSX Export ──────────────────────────────────────────────────────


@router.get("/export/{file_id}/xlsx")
async def export_xlsx(
    file_id: str,
    preset_id: str = Query(default="simple"),
    include: str = Query(default=""),
    exclude: str = Query(default=""),
    db: Session = Depends(get_db),
):
    """Preset-driven canonical XLSX export with metadata sheet."""

    boq_file, items = _get_file_with_items(file_id, db)

    # ── Load preset ──────────────────────────────────────────────
    preset = db.query(Preset).filter(Preset.id == preset_id).first()
    if not preset:
        raise HTTPException(
            status_code=404, detail=f"Preset '{preset_id}' not found"
        )

    include_cols = [c.strip() for c in include.split(",") if c.strip()]
    exclude_cols = [c.strip() for c in exclude.split(",") if c.strip()]
    columns = resolve_columns(preset.groups, include_cols or None, exclude_cols or None)
    num_cols = len(columns)

    now = datetime.now()
    now_str = now.strftime("%d.%m.%Y. %H:%M")
    project_name = boq_file.project_name or "N/A"
    grand_total = sum(i.total or 0 for i in items)

    # Check if status group is active
    status_group_active = any(c.group == "status" for c in columns)

    # ── Workbook ─────────────────────────────────────────────────
    wb = XLWorkbook()

    # ── Sheet 1: Troškovnik ──────────────────────────────────────
    ws = wb.active
    ws.title = "Tro\u0161kovnik"

    last_col_letter = get_column_letter(num_cols)
    merge_range = f"A{{row}}:{last_col_letter}{{row}}"

    # Row 1 — project title
    ws.merge_cells(merge_range.format(row=1))
    title_cell = ws["A1"]
    title_cell.value = project_name
    title_cell.font = Font(name="Arial", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — subtitle
    ws.merge_cells(merge_range.format(row=2))
    subtitle_cell = ws["A2"]
    subtitle_cell.value = (
        f"Izvor: {boq_file.file_name} | Preset: {preset.name} | Datum: {now_str}"
    )
    subtitle_cell.font = Font(
        name="Arial", size=9, italic=True, color="6b7280"
    )
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3 — spacer (blank)

    # Row 4 — column headers
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1a1a2e", end_color="1a1a2e", fill_type="solid"
    )
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    thin_border = Border(
        left=Side(style="thin", color="2a2a40"),
        right=Side(style="thin", color="2a2a40"),
        top=Side(style="thin", color="2a2a40"),
        bottom=Side(style="thin", color="2a2a40"),
    )

    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_def.label_hr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def.width

    # ── Data rows ────────────────────────────────────────────────
    section_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    ne_nudimo_fill = PatternFill(
        start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
    )

    data_start_row = 5
    for row_offset, item in enumerate(items):
        xl_row = data_start_row + row_offset
        is_section = item.item_type == "section_header"
        is_ne_nudimo = item.item_type == "ne_nudimo"
        is_composite = _is_composite_parent(item, items)

        for col_idx, col_def in enumerate(columns, start=1):
            value = _get_cell_value(item, col_def)

            # Handle image embedding for drawing column
            if col_def.col_type == "image" and value:
                file_path = str(value)
                if os.path.isfile(file_path):
                    try:
                        img = XLImage(file_path)
                        img.width = 80
                        img.height = 60
                        anchor = f"{get_column_letter(col_idx)}{xl_row}"
                        ws.add_image(img, anchor)
                        ws.row_dimensions[xl_row].height = 50
                    except Exception:
                        pass  # skip broken images
                    cell = ws.cell(row=xl_row, column=col_idx, value="")
                else:
                    cell = ws.cell(row=xl_row, column=col_idx, value="")
            else:
                cell = ws.cell(row=xl_row, column=col_idx, value=value)

            cell.border = thin_border

            # ── Column formatting ────────────────────────────
            if col_def.col_type == "float":
                cell.number_format = col_def.fmt
                cell.alignment = Alignment(
                    horizontal="right", vertical="center"
                )
            elif col_def.col_type in ("string", "text", "enum", "date", "image"):
                h_align = col_def.fmt if col_def.fmt in ("left", "center", "right") else "left"
                cell.alignment = Alignment(
                    horizontal=h_align,
                    vertical="center",
                    wrap_text=(col_def.col_type == "text"),
                )

            # ── Row-level styling ────────────────────────────
            if is_section:
                cell.font = Font(name="Arial", bold=True)
                cell.fill = section_fill
            elif is_ne_nudimo:
                cell.font = Font(name="Arial", strikethrough=True)
                cell.fill = ne_nudimo_fill
            elif is_composite:
                cell.font = Font(name="Arial", bold=True)

            # Status colour
            if col_def.key == "status" and value:
                status_color = _STATUS_COLORS.get(str(value))
                if status_color:
                    cell.fill = PatternFill(
                        start_color=status_color,
                        end_color=status_color,
                        fill_type="solid",
                    )

        # Section header: merge description across data columns
        if is_section and num_cols > 1:
            # Find the description column index
            desc_idx = None
            for ci, cd in enumerate(columns, start=1):
                if cd.key == "description":
                    desc_idx = ci
                    break
            if desc_idx and desc_idx < num_cols:
                ws.merge_cells(
                    start_row=xl_row,
                    start_column=desc_idx,
                    end_row=xl_row,
                    end_column=num_cols,
                )

    # ── Summary / UKUPNO row ─────────────────────────────────────
    if items:
        summary_row = data_start_row + len(items)
        top_border = Border(top=Side(style="medium", color="1a1a2e"))

        # Find the total column index
        total_col_idx = None
        for ci, cd in enumerate(columns, start=1):
            if cd.key == "total":
                total_col_idx = ci
                break

        if total_col_idx:
            # Put "UKUPNO:" label one column to the left of total, or in col 1
            label_col = max(total_col_idx - 1, 1)
            label_cell = ws.cell(
                row=summary_row, column=label_col, value="UKUPNO:"
            )
            label_cell.font = Font(name="Arial", bold=True)
            label_cell.border = top_border
            label_cell.alignment = Alignment(
                horizontal="right", vertical="center"
            )

            total_cell = ws.cell(
                row=summary_row, column=total_col_idx, value=grand_total
            )
            total_cell.font = Font(name="Arial", bold=True)
            total_cell.number_format = "#,##0.00"
            total_cell.alignment = Alignment(
                horizontal="right", vertical="center"
            )
            total_cell.border = top_border

            # Apply top border to remaining columns in summary row
            for ci in range(1, num_cols + 1):
                if ci not in (label_col, total_col_idx):
                    ws.cell(row=summary_row, column=ci).border = top_border

    # ── Sheet 2: Metadata ────────────────────────────────────────
    ws_meta = wb.create_sheet(title="Metadata")
    active_col_names = ", ".join(c.label_hr for c in columns)

    metadata_rows = [
        ("Klju\u010d", "Vrijednost"),
        ("Izvorni fajl", boq_file.file_name),
        ("Projekt", project_name),
        ("Preset", preset.name),
        ("Aktivni stupci", active_col_names),
        ("Datum izvoza", now_str),
        ("Verzija sheme", "1.0"),
        ("Broj stavki", str(len(items))),
        ("Ukupno", f"{grand_total:,.2f}"),
    ]

    meta_header_font = Font(name="Arial", bold=True)
    for row_idx, (key, val) in enumerate(metadata_rows, start=1):
        key_cell = ws_meta.cell(row=row_idx, column=1, value=key)
        val_cell = ws_meta.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            key_cell.font = meta_header_font
            val_cell.font = meta_header_font

    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 60

    # ── Sheet 3: Legenda (only if status group is active) ────────
    if status_group_active:
        ws_legend = wb.create_sheet(title="Legenda")
        ws_legend.cell(row=1, column=1, value="Status").font = Font(
            name="Arial", bold=True
        )
        ws_legend.cell(row=1, column=2, value="Opis").font = Font(
            name="Arial", bold=True
        )

        legend_items = [
            (ItemStatusEnum.PENDING, "Na \u010dekanju"),
            (ItemStatusEnum.ACCEPTED, "Prihva\u0107eno"),
            (ItemStatusEnum.REFUSED, "Odbijeno"),
            (ItemStatusEnum.NEGOTIATED, "Dogovoreno"),
            (ItemStatusEnum.EXPIRED, "Isteklo"),
        ]

        for idx, (status_enum, label) in enumerate(legend_items, start=2):
            status_cell = ws_legend.cell(row=idx, column=1, value=label)
            desc_cell = ws_legend.cell(row=idx, column=2, value=label)
            color = _STATUS_COLORS[status_enum.value]
            fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            status_cell.fill = fill
            desc_cell.fill = fill

        ws_legend.column_dimensions["A"].width = 16
        ws_legend.column_dimensions["B"].width = 20

    # ── Write to buffer & return ─────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    proj_part = _sanitize(project_name, 30)
    preset_part = _sanitize(preset.name, 20)
    date_part = now.strftime("%Y%m%d")
    filename = f"{proj_part}_{preset_part}_{date_part}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF Export ───────────────────────────────────────────────────────


@router.get("/export/{file_id}/pdf")
async def export_pdf(file_id: str, db: Session = Depends(get_db)):
    """Export a BOQ file as PDF."""
    boq_file, items = _get_file_with_items(file_id, db)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"BOQ Export - {boq_file.file_name}", styles["Title"]))
    elements.append(
        Paragraph(
            f"Project: {boq_file.project_name or 'N/A'} | "
            f"Items: {len(items)} | "
            f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 10 * mm))

    # Table
    header = ["#", "Description", "Unit", "Qty", "Unit Price", "Total"]
    data = [header]
    for item in items:
        desc = (item.description or "")[:80]  # truncate for PDF
        data.append([
            item.item_number or "",
            desc,
            item.unit or "",
            f"{item.quantity:.2f}" if item.quantity else "",
            f"{item.unit_price:.2f}" if item.unit_price else "",
            f"{item.total:.2f}" if item.total else "",
        ])

    # Summary
    grand_total = sum(i.total or 0 for i in items)
    data.append(["", "", "", "", "TOTAL:", f"{grand_total:.2f}"])

    col_widths = [25 * mm, 70 * mm, 15 * mm, 20 * mm, 25 * mm, 25 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            # Body
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (3, 1), (5, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Grid
            ("GRID", (0, 0), (-1, -2), 0.5, colors.HexColor("#2a2a40")),
            # Alternating rows
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8f8fc")]),
            # Summary row
            ("FONTNAME", (4, -1), (5, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#1a1a2e")),
            # Padding
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
    )
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    filename = boq_file.file_name.rsplit(".", 1)[0] + "_export.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
