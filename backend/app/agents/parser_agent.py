"""
Parser Agent - LLM-powered column detection for messy Excel BOQ files.

Uses a PydanticAI agent to analyze sample rows from Excel sheets and determine
which columns map to standard BOQ fields (itemNumber, description, unit,
quantity, unitPrice, total).
"""
from __future__ import annotations

from io import BytesIO
from typing import Optional

import openpyxl
from pydantic import BaseModel, Field
from pydantic_ai import Agent
from pydantic_ai.models.openai import OpenAIChatModel
from pydantic_ai.providers.openai import OpenAIProvider

from app.config import settings

# ---------------------------------------------------------------------------
# Pydantic models for structured output
# ---------------------------------------------------------------------------


class ColumnMapping(BaseModel):
    """Mapping of BOQ field names to zero-based column indices."""

    itemNumber: Optional[int] = Field(None, description="Column index for item/row number")
    description: Optional[int] = Field(None, description="Column index for item description")
    unit: Optional[int] = Field(None, description="Column index for unit of measure")
    quantity: Optional[int] = Field(None, description="Column index for quantity")
    unitPrice: Optional[int] = Field(None, description="Column index for unit price")
    total: Optional[int] = Field(None, description="Column index for total price")


class SheetMapping(BaseModel):
    """Column detection result for a single sheet."""

    headerRow: int = Field(
        description="Zero-based row index of the header row",
    )
    dataStartRow: int = Field(
        description="Zero-based row index where actual data begins (usually headerRow + 1)",
    )
    columns: ColumnMapping = Field(
        description="Mapping of BOQ field names to column indices",
    )
    skipPatterns: list[str] = Field(
        default_factory=list,
        description="Row patterns to skip (e.g. subtotal rows, section headers)",
    )


class SheetDetection(BaseModel):
    """Detection result for a single sheet in the workbook."""

    name: str = Field(description="Sheet name as it appears in the workbook")
    hasBOQData: bool = Field(
        description="Whether this sheet contains BOQ/cost data",
    )
    type: Optional[str] = Field(
        None,
        description=(
            "Sheet type: 'boq_data', 'summary', 'cover', 'conditions', "
            "'rekapitulacija', or null if unknown"
        ),
    )
    mapping: Optional[SheetMapping] = Field(
        None,
        description="Column mapping (only if hasBOQData is true)",
    )


class ColumnDetectionResult(BaseModel):
    """Complete column detection result for a workbook."""

    sheets: list[SheetDetection] = Field(
        description="Detection results for each analysed sheet",
    )


# ---------------------------------------------------------------------------
# LLM model & agent
# ---------------------------------------------------------------------------

_provider = OpenAIProvider(base_url=settings.LLM_BASE_URL)
_model = OpenAIChatModel(model_name=settings.LLM_MODEL_NAME, provider=_provider)

SYSTEM_PROMPT = """\
You are a specialist at reading messy Croatian construction BOQ (Bill of Quantities) \
Excel spreadsheets. Your job is to look at sample rows from each sheet and determine:

1. Whether the sheet contains BOQ pricing data (items with descriptions, units, \
quantities, prices) or is a cover page / summary / conditions sheet.
2. If it IS a BOQ data sheet, identify which column index (zero-based) corresponds \
to each of these standard fields:
   - itemNumber: The item or row number (e.g. "1.", "1.1", "A-01")
   - description: The textual description of the work item
   - unit: The unit of measure (m, m2, m3, kom, kg, kompl, pau\u0161, etc.)
   - quantity: The numeric quantity
   - unitPrice: The unit price (per-unit cost)
   - total: The total price (quantity * unitPrice)
3. Identify which row is the header row and where data starts.
4. Identify any row patterns that should be skipped (subtotals, section headers with \
no price data, "UKUPNO" rows, etc.).

Croatian BOQ files often have:
- Headers like: "R.br.", "Opis", "Jed.", "Kol.", "Jed. cijena", "Ukupno"
- Merged cells, empty rows, section headers
- Multiple header rows or repeated headers
- Columns shifted by one (empty column A)

Return a structured JSON result. If a column cannot be identified, omit it (null). \
For sheets that are NOT BOQ data, set hasBOQData to false and mapping to null.\
"""

parser_agent = Agent(
    _model,
    output_type=ColumnDetectionResult,
    system_prompt=SYSTEM_PROMPT,
    retries=2,
    model_settings={"temperature": 0.1},
)


# ---------------------------------------------------------------------------
# Helper: extract sample data from workbook
# ---------------------------------------------------------------------------

MAX_SAMPLE_ROWS = 30
MAX_SHEETS = 8
MAX_COLUMNS = 21  # up to column U


def _extract_sample_data(file_bytes: bytes, file_name: str) -> str:
    """Read an Excel file and return a text representation of sample rows per sheet."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames[:MAX_SHEETS]:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: \"{sheet_name}\" ===")

        row_count = 0
        for row_tuple in ws.iter_rows(
            min_row=1, max_col=MAX_COLUMNS, values_only=True,
        ):
            if row_count >= MAX_SAMPLE_ROWS:
                break
            cells = [
                str(cell)[:120] if cell is not None else ""
                for cell in row_tuple
            ]
            # Skip completely empty rows in sample
            if not any(c.strip() for c in cells):
                continue
            parts.append(
                f"  Row {row_count}: "
                + " | ".join(f"[{i}]{c}" for i, c in enumerate(cells) if c)
            )
            row_count += 1

        parts.append("")

    wb.close()
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

async def detect_columns_with_llm(
    file_bytes: bytes,
    file_name: str,
) -> ColumnDetectionResult:
    """Detect column mappings in a BOQ Excel file using the LLM.

    Parameters
    ----------
    file_bytes : bytes
        Raw Excel file content.
    file_name : str
        Original filename (used for context).

    Returns
    -------
    ColumnDetectionResult
        Structured detection result with per-sheet column mappings.
    """
    sample_text = _extract_sample_data(file_bytes, file_name)

    prompt = (
        f"Analyze this Excel file \"{file_name}\" and detect BOQ columns.\n\n"
        f"{sample_text}"
    )

    result = await parser_agent.run(prompt)
    return result.output
