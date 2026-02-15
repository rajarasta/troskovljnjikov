#!/usr/bin/env python3
"""Comprehensive BoQ Excel file analyzer."""
import os, re, sys
from collections import defaultdict, Counter

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

DIRECTORY = "/media/josip-rastocic/DrugiDisk/Programi/troskovljnjikov/vanjski-podaci/primjeri-excel-ponuda/"
SEP = "=" * 120


def trunc(val, m=60):
    s = str(val) if val is not None else ""
    return s[:m-3] + "..." if len(s) > m else s


def classify_a(val):
    if val is None:
        return "EMPTY"
    s = str(val).strip()
    if not s:
        return "EMPTY"
    if re.match(r'^[IVXLCDM]+\.?$', s):
        return "ROMAN"
    if re.match(r'^\d+(\.\d+)+\.?$', s):
        return "HIERARCHICAL"
    if re.match(r'^\d+\.?$', s):
        return "INTEGER"
    if re.match(r'^[A-Z](\.\d+)?\.?$', s):
        return "LETTER_CODE"
    try:
        float(s.replace(',', '.'))
        return "NUMERIC"
    except ValueError:
        pass
    return "TEXT(" + s[:30] + ")"


def detect_header(ws, mx=20):
    kws = {'opis', 'description', 'stavka', 'item', 'jed', 'jedinica', 'mj',
           'mjera', 'kolicina', 'kol', 'quantity', 'qty', 'cijena', 'price',
           'ukupno', 'total', 'iznos', 'amount', 'r.br', 'rbr', 'rb', 'red',
           'redni', 'broj', 'no', 'unit', 'jed.cijena', 'jedini', 'br',
           'pozicija', 'poz'}
    best_row, best_score = None, 0
    for ri in range(1, min(mx + 1, (ws.max_row or 0) + 1)):
        sc = 0
        for c in ws[ri]:
            if c.value:
                v = str(c.value).strip().lower()
                for k in kws:
                    if k in v:
                        sc += 1
                        break
        if sc > best_score:
            best_score = sc
            best_row = ri
    return best_row, best_score


def detect_boq_sheet(wb):
    kws = ['troskovnik', 'boq', 'bill', 'ponuda', 'offer', 'stavke',
           'items', 'sheet1', 'list1']
    scores = {}
    for n in wb.sheetnames:
        sc = sum(10 for k in kws if k in n.lower())
        ws = wb[n]
        if ws.max_row and ws.max_row > 5:
            sc += 2
        if ws.max_column and ws.max_column >= 4:
            sc += 1
        scores[n] = sc
    best = max(scores, key=scores.get) if scores else wb.sheetnames[0]
    return best, scores


def bold_analysis(ws, mx=200):
    bold, nonbold = [], []
    for ri in range(1, min(mx + 1, (ws.max_row or 0) + 1)):
        c = ws.cell(row=ri, column=2)
        if c.value and str(c.value).strip():
            if c.font and c.font.bold:
                bold.append((ri, trunc(c.value, 80)))
            else:
                nonbold.append((ri, trunc(c.value, 80)))
    return bold, nonbold


def find_formulas(ws):
    ff = []
    if not ws.max_column or not ws.max_row:
        return ff
    for ci in range(max(1, ws.max_column - 2), ws.max_column + 1):
        cl = get_column_letter(ci)
        for ri in range(1, min(ws.max_row + 1, 200)):
            c = ws.cell(row=ri, column=ci)
            if c.data_type == 'f' or (isinstance(c.value, str) and c.value.startswith('=')):
                ff.append((cl + str(ri), str(c.value)[:100]))
                if len(ff) >= 10:
                    return ff
    return ff


def row_heights(ws):
    dh = ws.sheet_format.defaultRowHeight or 15
    tall = []
    for ri in range(1, min((ws.max_row or 0) + 1, 500)):
        rd = ws.row_dimensions.get(ri)
        if rd and rd.height and rd.height > dh * 1.5:
            tall.append((ri, rd.height, trunc(ws.cell(row=ri, column=2).value, 60)))
    return dh, tall


def sample_rows(ws, n=5, after=None):
    samples = []
    start = (after + 1) if after else 1
    for ri in range(start, min((ws.max_row or 0) + 1, start + 200)):
        vals, empty = [], True
        for ci in range(1, min((ws.max_column or 0) + 1, 20)):
            v = ws.cell(row=ri, column=ci).value
            if v is not None and str(v).strip():
                empty = False
            vals.append(v)
        if not empty:
            samples.append((ri, vals))
        if len(samples) >= n:
            break
    return samples


def classify_family(info):
    cc = info.get('col_count', 0) or 0
    hs = info.get('header_score', 0)
    ap = info.get('col_a_patterns', {})
    hh = ap.get('HIERARCHICAL', 0) > 3
    hi = ap.get('INTEGER', 0) > 3
    tne = sum(v for k, v in ap.items() if k != 'EMPTY')
    mea = ap.get('EMPTY', 0) > tne * 2 if tne > 0 else True
    if cc <= 3:
        return "MINIMAL_COLS"
    if hs >= 4:
        if hh:
            return "STANDARD_HIERARCHICAL"
        if hi:
            return "STANDARD_SEQUENTIAL"
        return "STANDARD_MIXED"
    if hs >= 2:
        if mea:
            return "SPARSE_COL_A"
        return "PARTIAL_HEADERS"
    if hh or hi:
        return "NO_CLEAR_HEADERS"
    return "UNSTRUCTURED"


def analyze_file(fp):
    info = {'filename': os.path.basename(fp), 'error': None}
    try:
        wb = openpyxl.load_workbook(fp, read_only=False, data_only=False)
    except Exception as e:
        info['error'] = str(e)
        return info

    info['sheet_names'] = wb.sheetnames
    bsheet, sscores = detect_boq_sheet(wb)
    info['boq_sheet'] = bsheet
    info['sheet_scores'] = sscores
    ws = wb[bsheet]
    info['total_rows'] = ws.max_row
    info['col_count'] = ws.max_column

    hr, hsc = detect_header(ws)
    info['header_row'] = hr
    info['header_score'] = hsc
    hc = []
    if hr:
        for ci in range(1, min((ws.max_column or 0) + 1, 20)):
            hc.append(ws.cell(row=hr, column=ci).value)
    info['header_content'] = hc

    ac = Counter()
    for ri in range(1, min((ws.max_row or 0) + 1, 500)):
        ac[classify_a(ws.cell(row=ri, column=1).value)] += 1
    info['col_a_patterns'] = dict(ac)

    info['sample_rows'] = sample_rows(ws, 5, hr)

    br, nbr = bold_analysis(ws, 200)
    info['bold_rows'] = br[:15]
    info['bold_row_count'] = len(br)
    info['non_bold_row_count'] = len(nbr)

    info['merged_cells'] = [str(mr) for mr in ws.merged_cells.ranges]
    info['formulas'] = find_formulas(ws)

    dh, tr = row_heights(ws)
    info['default_row_height'] = dh
    info['tall_rows'] = tr[:10]

    info['format_family'] = classify_family(info)
    wb.close()
    return info


def print_analysis(info):
    print("")
    print(SEP)
    print("FILE: " + info['filename'])
    print(SEP)

    if info.get('error'):
        print("  *** ERROR: " + info['error'])
        return

    print("")
    print("  [1] SHEETS: " + str(info['sheet_names']))
    print("      Likely BoQ sheet: '" + info['boq_sheet'] + "'")
    print("      Sheet scores: " + str(info['sheet_scores']))

    print("")
    print("  [2] DIMENSIONS: " + str(info['total_rows']) + " rows x " + str(info['col_count']) + " columns")

    print("")
    print("  [3] HEADER ROW: row " + str(info['header_row']) + " (score: " + str(info['header_score']) + ")")
    for i, h in enumerate(info['header_content']):
        if h is not None:
            print("      " + get_column_letter(i + 1) + ": " + trunc(h, 80))

    print("")
    print("  [4] COLUMN A PATTERNS:")
    for p, c in sorted(info['col_a_patterns'].items(), key=lambda x: -x[1]):
        print("      " + str(p) + ": " + str(c))

    print("")
    print("  [5] SAMPLE ROWS (first 5 non-empty after header):")
    for ri, vals in info.get('sample_rows', []):
        dv = [trunc(v, 35) for v in vals]
        print("      Row " + str(ri).rjust(3) + ": " + str(dv))

    print("")
    print("  [6] BOLD IN COL B: Bold=" + str(info['bold_row_count']) +
          ", Non-bold=" + str(info['non_bold_row_count']))
    for ri, v in info['bold_rows'][:10]:
        print("        Bold Row " + str(ri).rjust(3) + ": " + v)

    mc = info.get('merged_cells', [])
    print("")
    print("  [7] MERGED CELLS: " + str(len(mc)) + " ranges")
    for m in mc[:15]:
        print("      " + m)
    if len(mc) > 15:
        print("      ... and " + str(len(mc) - 15) + " more")

    ff = info.get('formulas', [])
    print("")
    print("  [8] FORMULAS: " + str(len(ff)) + " found")
    for cr, f in ff:
        print("      " + cr + ": " + f)

    print("")
    print("  [9] ROW HEIGHTS (default: " + str(info['default_row_height']) + "pt):")
    tl = info.get('tall_rows', [])
    print("      Tall rows (>1.5x default): " + str(len(tl)))
    for ri, h, v in tl[:8]:
        print("        Row " + str(ri).rjust(3) + " (" + str(h) + "pt): " + v)

    print("")
    print("  [10] TOTAL ROWS: " + str(info['total_rows']))

    print("")
    print("  >>> FORMAT FAMILY: " + info['format_family'])


def main():
    print(SEP)
    print("COMPREHENSIVE BoQ EXCEL FILE ANALYSIS")
    print("Directory: " + DIRECTORY)
    print(SEP)

    files = sorted([
        os.path.join(DIRECTORY, f)
        for f in os.listdir(DIRECTORY)
        if f.endswith('.xlsx') and not f.startswith('~') and not f.startswith('.~')
    ])
    print("")
    print("Found " + str(len(files)) + " .xlsx files.")
    print("")

    all_info = []
    for fp in files:
        info = analyze_file(fp)
        all_info.append(info)
        print_analysis(info)

    hdr = '#' * 120

    # SUMMARY
    print("")
    print("")
    print(hdr)
    print("SUMMARY: FORMAT FAMILY GROUPING")
    print(hdr)

    fam = defaultdict(list)
    errs = []
    for i in all_info:
        if i.get('error'):
            errs.append(i['filename'])
        else:
            fam[i['format_family']].append(i['filename'])

    for f in sorted(fam):
        print("")
        print("  " + f)
        print("  " + '~' * len(f))
        for fn in fam[f]:
            print("    - " + fn)

    if errs:
        print("")
        print("  ERRORS")
        print("  " + '~' * 6)
        for f in errs:
            print("    - " + f)

    # CROSS-FILE STATS
    print("")
    print("")
    print(hdr)
    print("CROSS-FILE STATISTICS")
    print(hdr)

    ccs = [i['col_count'] for i in all_info if not i.get('error') and i.get('col_count')]
    rcs = [i['total_rows'] for i in all_info if not i.get('error') and i.get('total_rows')]
    hss = [i['header_score'] for i in all_info if not i.get('error')]

    if ccs:
        print("")
        print("  Column counts: min=" + str(min(ccs)) + ", max=" + str(max(ccs)) +
              ", values=" + str(sorted(set(ccs))))
    if rcs:
        print("  Row counts: min=" + str(min(rcs)) + ", max=" + str(max(rcs)) +
              ", avg=" + str(round(sum(rcs) / len(rcs))))
    if hss:
        print("  Header scores: min=" + str(min(hss)) + ", max=" + str(max(hss)) +
              ", avg=" + str(round(sum(hss) / len(hss), 1)))

    print("")
    print("  COMMON HEADER KEYWORDS:")
    hwc = Counter()
    for i in all_info:
        if i.get('error'):
            continue
        for h in i.get('header_content', []):
            if h:
                hwc[str(h).strip().lower()] += 1
    for w, c in hwc.most_common(25):
        print("    '" + w + "': " + str(c) + " files")

    print("")
    print("  COL A DOMINANT STYLE PER FILE:")
    ns = Counter()
    for i in all_info:
        if i.get('error'):
            continue
        ne = {k: v for k, v in i.get('col_a_patterns', {}).items() if k != 'EMPTY'}
        if ne:
            ns[max(ne.items(), key=lambda x: x[1])[0]] += 1
        else:
            ns['ALL_EMPTY'] += 1
    for s, c in ns.most_common():
        print("    " + s + ": " + str(c) + " files")

    print("")
    print("  EDGE CASES:")
    for i in all_info:
        if i.get('error'):
            print("    - " + i['filename'] + ": FAILED (" + i['error'][:80] + ")")
            continue
        issues = []
        if (i.get('col_count') or 0) <= 3:
            issues.append("Few cols (" + str(i['col_count']) + ")")
        if (i.get('col_count') or 0) > 12:
            issues.append("Many cols (" + str(i['col_count']) + ")")
        if i.get('header_score', 0) < 2:
            issues.append("Weak headers")
        if len(i.get('merged_cells', [])) > 20:
            issues.append("Heavy merging (" + str(len(i['merged_cells'])) + ")")
        if len(i.get('sheet_names', [])) > 3:
            issues.append("Many sheets (" + str(len(i['sheet_names'])) + ")")
        tr = i.get('total_rows') or 1
        if i.get('col_a_patterns', {}).get('EMPTY', 0) > tr * 0.8:
            issues.append("Col A mostly empty")
        if issues:
            print("    - " + i['filename'] + ": " + '; '.join(issues))

    print("")
    print(SEP)
    print("ANALYSIS COMPLETE")
    print(SEP)


if __name__ == '__main__':
    main()
